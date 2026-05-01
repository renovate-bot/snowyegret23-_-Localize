import argparse
import os
import struct
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCHEMAS: dict = {
    "StringTableParam.dat": {
        "entry_size": 0x108,
        "id_offset": 0x04,
        "text_fields": [(0x08, 0x100, "text")],
    },
    "BgmTable.dat": {
        "entry_size": 0x30,
        "id_offset": 0x00,
        "text_fields": [(0x04, 0x20, "title")],
    },
    "SeTable.dat": {
        "entry_size": 0x28,
        "id_offset": 0x00,
        "text_fields": [],
    },
    "CharacterTable.dat": {
        "entry_size": 0x2C,
        "id_offset": 0x00,
        "text_fields": [(0x0C, 0x10, "name"), (0x1C, 0x08, "alias")],
        # name: max 7자 (姓+名 형태), alias: max 3자 (이름 또는 역할표시 e.g. "男性")
    },
    "ScriptNameTable_spot.dat":        {"entry_size": 0x54, "id_offset": 0x00, "text_fields": [(0x14, 0x40, "title")]},
    "ScriptNameTable_alone.dat":       {"entry_size": 0x54, "id_offset": 0x00, "text_fields": [(0x14, 0x40, "title")]},
    "ScriptNameTable_camp.dat":        {"entry_size": 0x54, "id_offset": 0x00, "text_fields": [(0x14, 0x40, "title")]},
    "ScriptNameTable_chara.dat":       {"entry_size": 0x54, "id_offset": 0x00, "text_fields": [(0x14, 0x40, "title")]},
    "ScriptNameTable_randomEvent.dat": {"entry_size": 0x54, "id_offset": 0x00, "text_fields": [(0x14, 0x40, "title")]},
    "ScriptNameTable_article.dat":     {"entry_size": 0x54, "id_offset": 0x00, "text_fields": [(0x14, 0x40, "title")]},
    "SpotTableParam.dat": {
        "entry_size": 0x48,
        "id_offset": 0x00,
        "text_fields": [(0x10, 0x20, "name")],
    },
    "TouringTableParam.dat": {
        "entry_size": 0xA4,
        "id_offset": 0x00,
        "text_fields": [
            (0x22, 0x36, "road_name"),
            (0x62, 0x0E, "city_name"),
        ],
    },
    "GourmetSpotTableParam.dat": {
        "entry_size": 0x88,
        "id_offset": 0x00,
        "text_fields": [(0x08, 0x24, "name")],
    },
    "ArticleTableParam.dat": {
        "entry_size": 0x68,
        "id_offset": 0x00,
        "text_fields": [(0x08, 0x20, "title")],
    },
    "HeroineArticleTableParam.dat": {
        "entry_size": 0x38,
        "id_offset": 0x00,
        "text_fields": [(0x08, 0x20, "text")],
    },
    "TrendTable.dat": {
        "entry_size": 0x48,
        "id_offset": 0x00,
        "text_fields": [(0x04, 0x3C, "keyword")],
    },
    "ArticleScoreParam.dat":      {"entry_size": 0x24, "id_offset": None, "text_fields": []},
    "BgImageTable.dat":           {"entry_size": 0x3C, "id_offset": 0x00, "text_fields": []},
    "CharAnimeTable.dat":         {"entry_size": 0x14, "id_offset": 0x00, "text_fields": []},
    "CharImageTable.dat":         {"entry_size": 0x1C, "id_offset": 0x00, "text_fields": []},
    "ContestEntryTableParam.dat": {
        "entry_size": 0x40,
        "id_offset": 0x00,
        "text_fields": [(0x04, 0x20, "publisher"), (0x24, 0x10, "author")],
    },
    "ExifTableParam.dat":         {"entry_size": 0x34, "id_offset": 0x00, "text_fields": []},
    "HiddenRoute.dat":            {"entry_size": 0x30, "id_offset": 0x00, "text_fields": []},
    "IncomTableParam.dat":        {"entry_size": 0x24, "id_offset": 0x00, "text_fields": []},
    "PicTableParam.dat":          {"entry_size": 0x24, "id_offset": 0x00, "text_fields": []},
    "StillImageTable.dat":        {"entry_size": 0x40, "id_offset": 0x00, "text_fields": []},
    "StillTargetDirParam.dat":    {"entry_size": 0x1C, "id_offset": 0x00, "text_fields": []},
    "TextureIndexTable.dat":      {"entry_size": 0x2C, "id_offset": 0x00, "text_fields": []},
    "TownName.dat":               {"entry_size": 0x14, "id_offset": 0x00, "text_fields": []},
}


def voice_table_schema() -> dict:
    return {"entry_size": 0x14, "id_offset": 0x00, "text_fields": []}


def get_schema(filename: str) -> dict | None:
    if filename in SCHEMAS:
        return SCHEMAS[filename]
    if filename.startswith("VoiceTable_") and filename.endswith(".dat"):
        return voice_table_schema()
    return None


def read_utf16le(data: bytes, offset: int, max_bytes: int) -> str:
    chars: list[str] = []
    end = min(offset + max_bytes, len(data))
    i = offset
    while i < end - 1:
        cp = struct.unpack_from("<H", data, i)[0]
        if cp == 0:
            break
        if 0xD800 <= cp <= 0xDFFF:
            i += 2
            continue
        if cp < 0x20 and cp not in (0x09, 0x0A, 0x0D):
            break
        chars.append(chr(cp))
        i += 2
    return "".join(chars)


COLUMNS = ["table", "entry_idx", "entry_id", "field", "str_offset", "max_bytes", "original", "translation"]
COL_MAX_BYTES   = COLUMNS.index("max_bytes") + 1
COL_TRANSLATION = COLUMNS.index("translation") + 1


def _extract_rows(dat_path: str) -> list[dict]:
    fname = os.path.basename(dat_path)
    schema = get_schema(fname)
    if schema is None or not schema["text_fields"]:
        return []

    entry_size: int = schema["entry_size"]
    id_off = schema["id_offset"]
    fields = schema["text_fields"]

    data = open(dat_path, "rb").read()
    if len(data) % entry_size != 0:
        print(f"  [경고] {fname}: 파일 크기({len(data)})가 entry_size(0x{entry_size:X})로 나누어 떨어지지 않음")
        return []
    count = len(data) // entry_size
    table_key = os.path.splitext(fname)[0]

    rows: list[dict] = []
    for i in range(count):
        off = i * entry_size
        entry = data[off:off + entry_size]
        eid = struct.unpack_from("<I", entry, id_off)[0] if id_off is not None else i
        for f_off, f_max, f_name in fields:
            text = read_utf16le(entry, f_off, f_max)
            if not text.strip():
                continue
            rows.append({
                "table": table_key,
                "entry_idx": i,
                "entry_id": eid,
                "field": f_name,
                "str_offset": f_off,
                "max_bytes": f_max,
                "original": text,
                "translation": "",
            })
    return rows


def export_all(dat_dir: str, xlsx_path: str) -> None:
    parent = os.path.dirname(os.path.abspath(xlsx_path))
    if parent and not os.path.isdir(parent):
        os.makedirs(parent, exist_ok=True)

    all_rows: list[dict] = []
    per_table: list[tuple[str, int]] = []
    for fname in sorted(os.listdir(dat_dir)):
        if not fname.endswith(".dat"):
            continue
        rows = _extract_rows(os.path.join(dat_dir, fname))
        if rows:
            per_table.append((fname, len(rows)))
            all_rows.extend(rows)

    if not all_rows:
        print("추출할 문자열이 없습니다.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "strings"

    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for row in all_rows:
        ws.append([row[k] for k in COLUMNS])

    widths = {
        "A": 28,
        "B": 9,
        "C": 10,
        "D": 12,
        "E": 10,
        "F": 10,
        "G": 60,
        "H": 60,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    last_row = len(all_rows) + 1
    trans_col = get_column_letter(COL_TRANSLATION)
    max_col   = get_column_letter(COL_MAX_BYTES)
    rng = f"{trans_col}2:{trans_col}{last_row}"
    formula = f"LEN(${trans_col}2)*2+2>${max_col}2"
    yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=yellow))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    wb.save(xlsx_path)

    total = len(all_rows)
    for fname, n in per_table:
        print(f"  {fname}: {n} strings")
    print(f"완료: {total} strings -> {xlsx_path}")


def _read_xlsx_rows(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    out: list[dict] = []
    for row in rows_iter:
        if row is None or all(c is None or c == "" for c in row):
            continue
        rec = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
        out.append(rec)
    wb.close()
    return out


def import_all(xlsx_path: str, orig_dir: str, out_dir: str) -> None:
    os.makedirs(out_dir, exist_ok=True)

    rows = _read_xlsx_rows(xlsx_path)
    if not rows:
        print("XLSX 에 행이 없습니다.")
        return

    if "table" not in rows[0]:
        print(f"[오류] XLSX 첫 행에 'table' 컬럼이 없습니다. export 로 만든 포맷을 써주세요.")
        sys.exit(1)

    by_table: dict[str, list[dict]] = {}
    for r in rows:
        t = r.get("table")
        if not t:
            continue
        by_table.setdefault(str(t), []).append(r)

    all_overflows: list[str] = []
    pending_writes: list[tuple[str, bytes, int]] = []

    for table_key, t_rows in sorted(by_table.items()):
        dat_name = table_key + ".dat"
        orig_path = os.path.join(orig_dir, dat_name)
        if not os.path.exists(orig_path):
            print(f"  [경고] 원본 {dat_name} 없음 — 스킵")
            continue

        schema = get_schema(dat_name)
        if schema is None or not schema["text_fields"]:
            continue
        entry_size: int = schema["entry_size"]

        data = bytearray(open(orig_path, "rb").read())
        if len(data) % entry_size != 0:
            continue

        patches: list[tuple[int, int, bytes]] = []
        for row in t_rows:
            translation = row.get("translation")
            translation = (str(translation).strip() if translation is not None else "")
            if not translation:
                continue

            entry_idx = int(row["entry_idx"])
            str_off   = int(row["str_offset"])
            max_bytes = int(row["max_bytes"])
            field     = row.get("field") or "text"

            encoded = translation.encode("utf-16-le") + b"\x00\x00"
            if len(encoded) > max_bytes:
                limit_chars = max_bytes // 2 - 1
                all_overflows.append(
                    f"  {dat_name} entry[{entry_idx}].{field}: "
                    f"{len(translation)}자 > {limit_chars}자 한도 "
                    f"({len(encoded)}B > {max_bytes}B): \"{translation[:40]}...\""
                )
                continue

            abs_off = entry_idx * entry_size + str_off
            patches.append((abs_off, max_bytes, encoded))

        for abs_off, max_bytes, encoded in patches:
            data[abs_off:abs_off + max_bytes] = b"\x00" * max_bytes
            data[abs_off:abs_off + len(encoded)] = encoded
        pending_writes.append((dat_name, bytes(data), len(patches)))

    if all_overflows:
        print(f"[오류] 번역이 엔트리 한도를 초과한 행이 {len(all_overflows)} 개 있습니다.")
        print("      각 행을 한도 이내로 줄인 뒤 다시 import 하세요.")
        print("      (XLSX 의 translation 셀이 노란색이면 초과입니다)\n")
        for line in all_overflows:
            print(line)
        sys.exit(1)

    total = 0
    for dat_name, buf, n in pending_writes:
        open(os.path.join(out_dir, dat_name), "wb").write(buf)
        if n > 0:
            print(f"  -> {dat_name} ({n} patched)")
            total += n
    print(f"완료: {total} strings patched")


def dump_dat(dat_path: str) -> None:
    fname = os.path.basename(dat_path)
    schema = get_schema(fname)
    data = open(dat_path, "rb").read()
    if schema is None:
        print(f"{fname}: 스키마 없음 ({len(data):,}B)")
        return

    entry_size = schema["entry_size"]
    count = len(data) // entry_size
    print(f"{fname}: {count} entries x 0x{entry_size:X}")
    for i in range(count):
        off = i * entry_size
        entry = data[off:off + entry_size]
        eid_field = schema.get("id_offset")
        eid = struct.unpack_from("<I", entry, eid_field)[0] if eid_field is not None else i
        if not schema["text_fields"]:
            print(f"  [{i:4d}] id={eid}")
            continue
        texts = []
        for f_off, f_max, f_name in schema["text_fields"]:
            t = read_utf16le(entry, f_off, f_max)
            if t:
                texts.append(f"{f_name}=\"{t}\"")
        if texts:
            print(f"  [{i:4d}] id={eid}: " + " | ".join(texts))


def expand_string_table(orig_path: str, out_path: str, new_entry_size: int = 0x208) -> None:
    orig = open(orig_path, "rb").read()
    old_size = 0x108
    if len(orig) % old_size != 0:
        raise SystemExit(f"입력 파일이 0x{old_size:X} 배수가 아니다")
    count = len(orig) // old_size

    if new_entry_size < old_size:
        raise SystemExit("new_entry_size 는 0x108 이상이어야 한다")

    out = bytearray(count * new_entry_size)
    for i in range(count):
        src = orig[i * old_size : (i + 1) * old_size]
        dst_off = i * new_entry_size
        out[dst_off : dst_off + old_size] = src

    open(out_path, "wb").write(out)
    print(f"확장 완료: {count} entries, 엔트리 크기 0x{old_size:X} → 0x{new_entry_size:X}")
    print(f"주의: Furaiki5_steam.exe 내 0x108 상수 패치가 선행되어야 정상 동작한다.")


def merge_xlsx(src_xlsx: str, dst_xlsx: str, key: str = "auto") -> None:
    """기존 (구) XLSX의 translation 열을 새(현재) XLSX로 옮긴다.

    매핑 규칙:
      key='auto'    : 가능하면 (table, original) 기준, 없으면 (table, entry_idx, field) fallback
      key='text'    : (table, original) 만 사용
      key='index'   : (table, entry_idx, field) 만 사용

    스키마가 바뀌어 entry_idx가 달라진 경우에도 원본 일본어 텍스트가 같으면 매핑이 유지된다.
    동일 (table, original)이 src에 여러 번 나오면 비어있지 않은 마지막 값을 채택한다.
    """
    src_wb = load_workbook(src_xlsx, data_only=True, read_only=True)
    src_ws = src_wb.active
    src_rows = src_ws.iter_rows(values_only=True)
    src_hdr = [str(c) if c is not None else "" for c in next(src_rows)]
    si_table = src_hdr.index("table") if "table" in src_hdr else None
    si_idx   = src_hdr.index("entry_idx") if "entry_idx" in src_hdr else None
    si_field = src_hdr.index("field") if "field" in src_hdr else None
    si_orig  = src_hdr.index("original") if "original" in src_hdr else None
    si_trans = src_hdr.index("translation") if "translation" in src_hdr else None
    if si_table is None or si_trans is None:
        raise SystemExit("[오류] src XLSX 에 'table' 또는 'translation' 컬럼이 없습니다.")

    by_text: dict = {}   # (table, field, original) -> translation
    by_text_t: dict = {} # (table, original) -> translation  (field 무시 fallback)
    by_idx:  dict = {}   # (table, entry_idx, field) -> translation
    n_src = 0
    for row in src_rows:
        if row is None or all(c is None or c == "" for c in row): continue
        t = row[si_table] if si_table is not None else None
        tr = row[si_trans] if si_trans is not None else None
        if t is None or tr is None or str(tr).strip() == "":
            continue
        tr = str(tr)
        t = str(t)
        f = (str(row[si_field]) if si_field is not None and row[si_field] is not None else None)
        o = (str(row[si_orig]) if si_orig is not None and row[si_orig] is not None else None)
        i = (int(row[si_idx]) if si_idx is not None and row[si_idx] is not None else None)
        if o is not None:
            by_text[(t, f, o)] = tr
            by_text_t[(t, o)] = tr
        if i is not None and f is not None:
            by_idx[(t, i, f)] = tr
        n_src += 1
    src_wb.close()
    print(f"src에서 {n_src}개 번역 로드 (text-key {len(by_text)}, idx-key {len(by_idx)})")

    dst_wb = load_workbook(dst_xlsx)
    dst_ws = dst_wb.active
    dst_hdr = [dst_ws.cell(1, c).value for c in range(1, dst_ws.max_column + 1)]
    dt_table = dst_hdr.index("table") + 1
    dt_idx   = dst_hdr.index("entry_idx") + 1
    dt_field = dst_hdr.index("field") + 1
    dt_orig  = dst_hdr.index("original") + 1
    dt_trans = dst_hdr.index("translation") + 1

    n_text=n_text_t=n_idx=n_skip=n_overwrite=0
    for r in range(2, dst_ws.max_row + 1):
        t = dst_ws.cell(r, dt_table).value
        if t is None: continue
        t = str(t)
        existing = dst_ws.cell(r, dt_trans).value
        f = dst_ws.cell(r, dt_field).value
        f = str(f) if f is not None else None
        o = dst_ws.cell(r, dt_orig).value
        o = str(o) if o is not None else None
        i = dst_ws.cell(r, dt_idx).value
        i = int(i) if i is not None else None

        kr = None
        if key in ("auto", "text"):
            if o is not None and (t, f, o) in by_text:
                kr = by_text[(t, f, o)]; n_text += 1
            elif o is not None and (t, o) in by_text_t:
                # field 이름이 바뀐 경우 (예: ContestEntry "entry" → "publisher")
                kr = by_text_t[(t, o)]; n_text_t += 1
        if kr is None and key in ("auto", "index"):
            if i is not None and f is not None and (t, i, f) in by_idx:
                kr = by_idx[(t, i, f)]; n_idx += 1
        if kr is None:
            n_skip += 1
            continue
        if existing and str(existing).strip() and str(existing) != kr:
            n_overwrite += 1
        dst_ws.cell(r, dt_trans).value = kr

    dst_wb.save(dst_xlsx)
    print(f"merge 완료: text-match {n_text}, text-only-match {n_text_t}, idx-match {n_idx}, "
          f"기존값 덮어씀 {n_overwrite}, 매칭실패 {n_skip}")


def main() -> None:
    p = argparse.ArgumentParser(description="풍우래기5 .dat 스키마 도구")
    sub = p.add_subparsers(dest="cmd", required=True)

    pe = sub.add_parser("export", help=".dat → 단일 XLSX 추출")
    pe.add_argument("dat_dir"); pe.add_argument("xlsx_path")

    pi = sub.add_parser("import", help="XLSX → .dat 리패킹")
    pi.add_argument("xlsx_path"); pi.add_argument("orig_dir"); pi.add_argument("output_dir")

    pd = sub.add_parser("dump", help="단일 .dat 덤프")
    pd.add_argument("file")

    px = sub.add_parser("expand", help="StringTableParam.dat 엔트리 크기 확장 (실험)")
    px.add_argument("orig"); px.add_argument("out")
    px.add_argument("--size", type=lambda s: int(s, 0), default=0x208,
                    help="새 엔트리 크기 (기본 0x208)")

    pm = sub.add_parser("merge", help="기존 XLSX → 새 XLSX 번역 병합 (스키마 변경시 사용)")
    pm.add_argument("src_xlsx", help="기존(구버전) XLSX")
    pm.add_argument("dst_xlsx", help="새(현재 스키마) XLSX (in-place 수정됨)")
    pm.add_argument("--key", choices=["auto", "text", "index"], default="auto",
                    help="매칭 키: auto(원본텍스트 우선)/text(원본만)/index(인덱스만) (기본 auto)")

    args = p.parse_args()
    if args.cmd == "export":
        export_all(args.dat_dir, args.xlsx_path)
    elif args.cmd == "import":
        import_all(args.xlsx_path, args.orig_dir, args.output_dir)
    elif args.cmd == "dump":
        dump_dat(args.file)
    elif args.cmd == "expand":
        expand_string_table(args.orig, args.out, args.size)
    elif args.cmd == "merge":
        merge_xlsx(args.src_xlsx, args.dst_xlsx, args.key)


if __name__ == "__main__":
    main()
